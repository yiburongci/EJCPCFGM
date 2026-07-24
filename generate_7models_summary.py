"""Summarize predictions of 8 forecasting models into one Excel file.

Models: EJCPCFGM, ARIMA, GM(1,1), FGM, JFGM, NGBM, LSSVR (best), LSTM (best)
Datasets: 4 (Jiangsu/Anhui x Elec/Gen). Train 2010-2021, Test 2022-2024.
"""

import os
import csv
import warnings
import numpy as np
from scipy.optimize import minimize_scalar, minimize
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

warnings.filterwarnings('ignore')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LSTM_DIR = os.path.join(SCRIPT_DIR, 'Grid_Search_LSTM')
LSSVR_DIR = os.path.join(SCRIPT_DIR, 'Grid_Search_LSSVR')

# ==================== 数据定义 ====================
DATASETS = {
    'Anhui_Consumption': {
        'train': [1078.00, 1221.19, 1361.10, 1528.10, 1585.18, 1639.79,
                   1794.98, 1921.48, 2135.07, 2301.00, 2428.00, 2715.00],
        'test':  [2993.00, 3214.00, 3597.86],
        'train_years': list(range(2010, 2022)),
        'test_years': [2022, 2023, 2024],
    },
    'Jiangsu_Consumption': {
        'train': [3864.00, 4281.62, 4580.90, 4956.60, 5012.54, 5114.70,
                   5458.95, 5807.89, 6128.27, 6264.00, 6374.00, 7101.00],
        'test':  [7400.00, 7833.00, 8486.93],
        'train_years': list(range(2010, 2022)),
        'test_years': [2022, 2023, 2024],
    },
    'Anhui_Generation': {
        'train': [1443.85, 1635.35, 1767.50, 1970.04, 2033.91, 2062.00,
                   2252.69, 2456.28, 2734.49, 2886.67, 2808.98, 3083.39],
        'test':  [3298.77, 3549.45, 3863.46],
        'train_years': list(range(2010, 2022)),
        'test_years': [2022, 2023, 2024],
    },
    'Jiangsu_Generation': {
        'train': [3359.18, 3762.50, 3928.40, 4320.68, 4347.57, 4361.00,
                   4709.37, 4914.74, 5085.08, 5166.43, 5217.54, 5968.89],
        'test':  [6077.31, 6390.53, 6807.37],
        'train_years': list(range(2010, 2022)),
        'test_years': [2022, 2023, 2024],
    },
}

# ==================== LSTM 最优配置（直接读取 LSTM_gridsearch.py 输出）====================
LSTM_BEST = {
    'Anhui_Consumption':  {'ws': 4, 'hidden': 128, 'layers': 1, 'lr': 0.01},
    'Jiangsu_Consumption': {'ws': 6, 'hidden': 128, 'layers': 2, 'lr': 0.001},
    'Anhui_Generation':    {'ws': 6, 'hidden': 64, 'layers': 1, 'lr': 0.001},
    'Jiangsu_Generation': {'ws': 5, 'hidden': 128, 'layers': 2, 'lr': 0.001},
}


def load_lstm_results(ds_name, train_data, test_data):
    pred_csv = os.path.join(LSTM_DIR, f'{ds_name}_predictions.csv')
    if not os.path.exists(pred_csv):
        return None, None, None

    fitted_list, forecast_list = [], []
    ws_offset = None

    with open(pred_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['Type'] == 'Train':
                if ws_offset is None:
                    ws_offset = int(row['Year_Index'])
                fitted_list.append(float(row['Predicted']))
            else:
                forecast_list.append(float(row['Predicted']))

    if not fitted_list or not forecast_list:
        return None, None, None
    return np.array(fitted_list), np.array(forecast_list), ws_offset


def load_lssvr_results(ds_name):
    pred_csv = os.path.join(LSSVR_DIR, f'{ds_name}_predictions.csv')
    if not os.path.exists(pred_csv):
        return None, None, None

    fitted_list, forecast_list = [], []
    ws_offset = None

    with open(pred_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['Type'] == 'Train':
                if ws_offset is None:
                    ws_offset = int(row['Year_Index'])
                fitted_list.append(float(row['Predicted']))
            else:
                forecast_list.append(float(row['Predicted']))

    if not fitted_list or not forecast_list:
        return None, None, None
    return np.array(fitted_list), np.array(forecast_list), ws_offset


# ==================== EJCPCFGM 最优配置 ====================
# 来自各省份独立脚本的最优参数
EJCPCFGM_BEST = {
    'Anhui_Consumption': {
        'seed': 94439, 'n_pulses': 3,
        'ridge_lambda': 1e-9, 'theta_lb': -5.0, 'cq_ub': 3.5,
    },
    'Jiangsu_Consumption': {
        'seed': 29165, 'n_pulses': 3,
        'ridge_lambda': 1e-4, 'theta_lb': -5.0, 'cq_ub': 2.5,
    },
    'Anhui_Generation': {
        'seed': 88199, 'n_pulses': 3,
        'ridge_lambda': 1e-6, 'theta_lb': -4.5, 'cq_ub': 1.5,
    },
    'Jiangsu_Generation': {
        'seed': 51501, 'n_pulses': 3,
        'ridge_lambda': 1e-9, 'theta_lb': -6.0, 'cq_ub': 3.0,
    },
}


# ==================== 工具函数 ====================
def safe_div(a, b):
    return a / b if abs(b) > 1e-10 else 0.0


def calculate_mape(actual, predicted):
    actual = np.array(actual, dtype=float)
    predicted = np.array(predicted, dtype=float)
    mask = actual != 0
    if np.sum(mask) == 0:
        return 0.0
    return float(np.mean(np.abs((actual[mask] - predicted[mask]) / actual[mask])) * 100)


# ==================== 1. GM(1,1) ====================
def gm11_predict(train_data, n_forecast=3):
    """
    GM(1,1) 模型
    1-AGO: x1(k) = sum(x0(1:k))
    白化方程: dx1/dt + a*x1 = b
    时间响应: x1_hat(k) = (x0(1) - b/a)*exp(-a*(k-1)) + b/a
    IAGO: x0_hat(k) = x1_hat(k) - x1_hat(k-1)
    """
    try:
        x0 = np.array(train_data, dtype=float)
        n = len(x0)
        x1 = np.cumsum(x0)
        z1 = 0.5 * x1[1:] + 0.5 * x1[:-1]
        B = np.vstack((-z1, np.ones(n - 1))).T
        Y = x0[1:]
        u = np.linalg.lstsq(B, Y, rcond=None)[0]
        a, b = u[0], u[1]

        def x1_hat_func(k_val):
            return (x0[0] - safe_div(b, a)) * np.exp(-a * (k_val - 1)) + safe_div(b, a)

        total_steps = n + n_forecast
        x1_all = np.array([x1_hat_func(k + 1) for k in range(total_steps)])

        x0_all = np.zeros(total_steps)
        x0_all[0] = x1_all[0]
        for k in range(1, total_steps):
            x0_all[k] = x1_all[k] - x1_all[k - 1]

        return {
            'fitted': x0_all[:n],
            'forecast': x0_all[n:],
            'a': a, 'b': b,
            'success': True
        }
    except Exception:
        return {'fitted': None, 'forecast': None, 'success': False}


# ==================== 2. FGM 分数阶灰预测模型 ====================
def fgm_predict(train_data, n_forecast=3, r=None):
    """
    FGM(r, 1) 分数阶累加灰预测模型

    分数阶累加 (FAGO):
    x^(r)(k) = sum_{i=1}^k C(r+k-i-1, k-i) * x^(0)(i)

    白化方程: dx^(r)/dt + a*x^(r) = b

    严格初始条件反解:
    令 x^(r)(1) = x^(0)(1)
    x^(r)_hat(k) = (x^(0)(1) - b/a) * exp(-a*(k-1)) + b/a

    逆累加还原 (IAGO):
    x^(0)_hat(k) = sum_{i=1}^k C(-r+k-i-1, k-i) * x^(r)_hat(i)
    """
    try:
        from scipy.special import gamma

        x0 = np.array(train_data, dtype=float)
        n = len(x0)

        def frac_comb(r_val, k):
            if k == 0:
                return 1.0
            try:
                return gamma(r_val + k) / (gamma(k + 1) * gamma(r_val))
            except:
                return 0.0

        def fago_seq(x, r_val):
            result = np.zeros(len(x))
            for k in range(len(x)):
                s = sum(frac_comb(r_val, k - i) * x[i] for i in range(k + 1))
                result[k] = s
            return result

        def ifago_seq(x, r_val):
            result = np.zeros(len(x))
            for k in range(len(x)):
                s = sum(frac_comb(-r_val, k - i) * x[i] for i in range(k + 1))
                result[k] = s
            return result

        def objective(r_val):
            if r_val <= 0.01 or r_val > 2.0:
                return 1e10
            xr = fago_seq(x0, r_val)
            zr = 0.5 * xr[1:] + 0.5 * xr[:-1]
            B = np.vstack((-zr, np.ones(n - 1))).T
            Y = xr[1:] - xr[:-1]
            try:
                u = np.linalg.lstsq(B, Y, rcond=None)[0]
                a, b = u[0], u[1]
            except:
                return 1e10

            def x1_hat_func(k_val):
                return (x0[0] - safe_div(b, a)) * np.exp(-a * (k_val - 1)) + safe_div(b, a)

            x1_all = np.array([x1_hat_func(k + 1) for k in range(n)])
            x0_hat = ifago_seq(x1_all, r_val)
            mape = np.mean(np.abs((x0 - x0_hat) / x0)) * 100
            return mape if np.isfinite(mape) else 1e10

        if r is None:
            res = minimize_scalar(objective, bounds=(0.01, 2.0), method='bounded',
                                 options={'xatol': 0.001})
            r_opt = res.x
        else:
            r_opt = r

        xr = fago_seq(x0, r_opt)
        zr = 0.5 * xr[1:] + 0.5 * xr[:-1]
        B = np.vstack((-zr, np.ones(n - 1))).T
        Y = xr[1:] - xr[:-1]
        u = np.linalg.lstsq(B, Y, rcond=None)[0]
        a, b = u[0], u[1]

        def x1_hat_func(k_val):
            return (x0[0] - safe_div(b, a)) * np.exp(-a * (k_val - 1)) + safe_div(b, a)

        total_steps = n + n_forecast
        x1_all = np.array([x1_hat_func(k + 1) for k in range(total_steps)])
        x0_all = ifago_seq(x1_all, r_opt)

        return {
            'fitted': x0_all[:n],
            'forecast': x0_all[n:],
            'r': r_opt, 'a': a, 'b': b,
            'success': True
        }
    except Exception:
        return {'fitted': None, 'forecast': None, 'success': False}


# ==================== 3. JFGM 跳跃分数阶灰预测模型 ====================
def jfgm_predict(train_data, n_forecast=3, r=None, S=None, ad=None, t0_idx=None):
    """
    JFGM Jump Fractional Grey Model

    分数阶累加 (FAGO): xr(k) = sum_{i=1}^k C(r+k-i-1, k-i) * x0(i)
    白化方程: dx^(r)/dt + a*x^(r) = b + c*S*u(t-t0)

    分段时间响应:
    k < t0:  xr_hat(k) = (x0(1) - b/a) * exp(-a*(k-1)) + b/a
    k >= t0: xr_hat(k) = (x0(1) - b/a - c/a) * exp(-a*(k-1)) + c*ad^(k-t0)/a + b/a

    优化参数: r (分数阶), S (冲击强度), ad (冲击衰减)
    """
    try:
        from scipy.optimize import minimize

        x0 = np.array(train_data, dtype=float)
        n = len(train_data)

        if t0_idx is None:
            # 2020年末/2021年初产生冲击
            # k=11对应2020年, k=12对应2021年
            # B矩阵有n-1行(k=2到k=n), 要让k=12受影响, 条件 k>=12
            t0_idx = n - 1  # = 11, 条件 k >= 12 时有冲击

        def frac_comb(r_val, k):
            if k == 0:
                return 1.0
            try:
                from scipy.special import gammaln
                from math import exp
                log_val = gammaln(r_val + k) - gammaln(k + 1) - gammaln(r_val)
                return exp(log_val)
            except:
                return 0.0

        def fago_seq(x, r_val):
            result = np.zeros(len(x))
            for k in range(len(x)):
                result[k] = sum(frac_comb(r_val, k - i) * x[i] for i in range(k + 1))
            return result

        def objective(params):
            r_opt, S_opt, ad_opt = params[0], params[1], params[2]
            if r_opt < 0.001 or r_opt > 2.0 or S_opt < 0.001 or S_opt > 10 or ad_opt < 0.1 or ad_opt > 2.0:
                return 1e9
            try:
                xr = fago_seq(x0, r_opt)
                if not np.all(np.isfinite(xr)):
                    return 1e9
                z = 0.5 * xr[1:] + 0.5 * xr[:-1]
                Y = xr[1:] - xr[:-1]
                B = np.zeros((n - 1, 3))
                B[:, 0] = -z
                for i in range(n - 1):
                    k = i + 2  # 1-indexed, k=2 to n
                    if k >= t0_idx + 1:  # k >= t0 时有冲击
                        B[i, 1] = S_opt
                    else:
                        B[i, 1] = 0.0
                B[:, 2] = 1.0
                u = np.linalg.inv(B.T @ B) @ B.T @ Y
                a, c_shock, b = u[0], u[1], u[2]
                if not np.isfinite(a) or not np.isfinite(c_shock) or not np.isfinite(b):
                    return 1e9
            except Exception as e:
                return 1e9

            try:
                xr_hat = np.zeros(n)
                xr_hat[0] = x0[0]
                for i in range(1, n):
                    k_paper = i + 1
                    t0_paper = t0_idx + 1  # = 12
                    if k_paper < t0_paper:
                        xr_hat[i] = (x0[0] - b / a) * np.exp(-a * (k_paper - 1)) + b / a
                    else:
                        shock_term = (b + c_shock * (ad_opt ** (k_paper - t0_paper))) / a
                        xr_hat[i] = (x0[0] - shock_term) * np.exp(-a * (k_paper - 1)) + shock_term
                x0_hat = fago_seq(xr_hat, -r_opt)
                if not np.all(np.isfinite(x0_hat)):
                    return 1e9
                mape = np.mean(np.abs((x0 - x0_hat) / x0))
                if not np.isfinite(mape):
                    return 1e9
                return mape
            except Exception as e:
                return 1e9

        if r is None or S is None or ad is None:
            from scipy.optimize import minimize

            best_result = None
            best_loss = float('inf')

            # Multi-start optimization over (r, S, ad)
            for init_r, init_S, init_ad in [
                (0.3, 2.0, 1.0), (0.5, 5.0, 1.0), (0.8, 1.0, 1.0),
                (1.0, 3.0, 1.0), (1.5, 8.0, 1.0), (0.5, 0.5, 0.5),
                (1.0, 0.5, 1.5), (0.3, 5.0, 0.5), (0.8, 3.0, 1.5),
                (1.2, 1.0, 0.8),
            ]:
                try:
                    opt_result = minimize(
                        objective, [init_r, init_S, init_ad],
                        method='Nelder-Mead',
                        options={'xatol': 1e-8, 'fatol': 1e-6, 'maxiter': 3000}
                    )
                    if opt_result.fun < best_loss:
                        best_loss = opt_result.fun
                        best_result = opt_result
                except:
                    continue

            if best_result is None or best_loss >= 1.0:
                return {'fitted': None, 'forecast': None, 'success': False}

            r_opt, S_opt, ad_opt = best_result.x[0], best_result.x[1], best_result.x[2]
        else:
            r_opt, S_opt = r, S
            ad_opt = ad

        xr = fago_seq(x0, r_opt)
        z = 0.5 * xr[1:] + 0.5 * xr[:-1]
        Y = xr[1:] - xr[:-1]
        B = np.zeros((n - 1, 3))
        B[:, 0] = -z
        for i in range(n - 1):
            k = i + 2
            if k >= t0_idx + 1:  # k >= t0 时有冲击
                B[i, 1] = S_opt
            else:
                B[i, 1] = 0.0
        B[:, 2] = 1.0
        beta = np.linalg.inv(B.T @ B) @ B.T @ Y
        a, c_shock, b = beta[0], beta[1], beta[2]


        total_len = n + n_forecast
        xr_hat = np.zeros(total_len)
        xr_hat[0] = x0[0]
        for i in range(1, total_len):
            k_paper = i + 1
            t0_paper = t0_idx + 1
            if k_paper < t0_paper:
                xr_hat[i] = (x0[0] - b / a) * np.exp(-a * (k_paper - 1)) + b / a
            else:
                shock_term = (b + c_shock * (ad_opt ** (k_paper - t0_paper))) / a
                xr_hat[i] = (x0[0] - shock_term) * np.exp(-a * (k_paper - 1)) + shock_term
        x0_hat = fago_seq(xr_hat, -r_opt)

        return {
            'fitted': x0_hat[:n],
            'forecast': x0_hat[n:],
            'r': r_opt, 'S': S_opt, 'ad': ad_opt,
            'success': True
        }
    except Exception as e:
        import traceback
        return {'fitted': None, 'forecast': None, 'success': False}


# ==================== 4. NGBM 非线性灰伯努利模型 ====================
def ngbm_predict(train_data, n_forecast=3, n_exp=None):
    """
    NGBM Nonlinear Grey Bernoulli Model (verified implementation from 7methods_comparison.py)
    
    微分方程: dx^1/dt + a*x^1 = b*(x^1)^n
    
    通过 y = (x^1)^(1-n) 变换后:
    时间响应: x^1_hat(k) = [(x^0(1)^(1-n) - b/a) * exp(-a*(1-n)*k) + b/a]^(1/(1-n))
    IAGO: x^0_hat(k) = x^1_hat(k) - x^1_hat(k-1)
    """
    try:
        from scipy.optimize import minimize_scalar

        x0 = np.array(train_data, dtype=float)
        m = len(x0)

        def objective(n_opt):
            if np.isclose(n_opt, 1.0) or n_opt < -2 or n_opt > 2:
                return 1e9
            x1 = np.cumsum(x0)
            z1 = 0.5 * x1[1:] + 0.5 * x1[:-1]
            B = np.zeros((m - 1, 2))
            B[:, 0] = -z1
            B[:, 1] = z1 ** n_opt
            Y_N = x0[1:]
            try:
                params = np.linalg.inv(B.T @ B) @ B.T @ Y_N
                a, b = params[0], params[1]
            except:
                return 1e9
            
            x1_hat = np.zeros(m)
            x1_hat[0] = x0[0]
            for k in range(1, m):
                base = (x0[0] ** (1 - n_opt) - b / a) * np.exp(-a * (1 - n_opt) * k) + b / a
                if base < 0:
                    return 1e9
                x1_hat[k] = base ** (1 / (1 - n_opt))
            
            x0_hat = np.zeros(m)
            x0_hat[0] = x1_hat[0]
            for k in range(1, m):
                x0_hat[k] = x1_hat[k] - x1_hat[k - 1]
            
            arpe = np.mean(np.abs((x0[1:] - x0_hat[1:]) / x0[1:])) * 100
            return arpe if np.isfinite(arpe) else 1e9

        if n_exp is None:
            opt_result = minimize_scalar(objective, bounds=(-2.0, 2.0), method='bounded')
            n_opt = opt_result.x
        else:
            n_opt = n_exp

        x1 = np.cumsum(x0)
        z1 = 0.5 * x1[1:] + 0.5 * x1[:-1]
        B = np.zeros((m - 1, 2))
        B[:, 0] = -z1
        B[:, 1] = z1 ** n_opt
        Y_N = x0[1:]
        params = np.linalg.inv(B.T @ B) @ B.T @ Y_N
        a, b = params[0], params[1]

        total_len = m + n_forecast
        x1_hat = np.zeros(total_len)
        x1_hat[0] = x0[0]
        for k in range(1, total_len):
            base = (x0[0] ** (1 - n_opt) - b / a) * np.exp(-a * (1 - n_opt) * k) + b / a
            if base < 0:
                return {'fitted': None, 'forecast': None, 'success': False, 'error': 'overflow'}
            x1_hat[k] = base ** (1 / (1 - n_opt))

        x0_hat = np.zeros(total_len)
        x0_hat[0] = x1_hat[0]
        for k in range(1, total_len):
            x0_hat[k] = x1_hat[k] - x1_hat[k - 1]

        return {
            'fitted': x0_hat[:m],
            'forecast': x0_hat[m:],
            'n': n_opt, 'a': a, 'b': b,
            'success': True
        }
    except Exception:
        return {'fitted': None, 'forecast': None, 'success': False}


# ==================== 5. ARIMA ====================
def arima_predict(train_data, n_forecast=3, order=(1, 1, 1)):
    try:
        from statsmodels.tsa.arima.model import ARIMA
        model = ARIMA(train_data, order=order)
        model_fit = model.fit()
        fitted_vals = model_fit.fittedvalues
        predictions = model_fit.forecast(steps=n_forecast)

        fitted = np.zeros(len(train_data))
        fitted[0] = train_data[0]
        n_skip = len(train_data) - len(fitted_vals)
        if n_skip > 0 and len(fitted_vals) > 0:
            fitted[n_skip:] = fitted_vals[:len(train_data) - n_skip]
        elif len(fitted_vals) > 0:
            fitted = fitted_vals[-len(train_data):]

        return {
            'fitted': fitted,
            'forecast': np.array(predictions),
            'success': True
        }
    except Exception:
        return {'fitted': None, 'forecast': None, 'success': False}


# ==================== 7. EJCPCFGM (EJCPEJCPCFGM) 模型 ====================
def cfgm_predict(train_data, test_data, cfg):
    """
    EJCPCFGM (EJCPEJCPCFGM) 模型 - 来自各省份独立脚本的最优配置
    使用 run_ejcpcfgm 函数进行预测
    """
    try:
        np.random.seed(cfg['seed'])
        import ejcpcfgm_model
        ejcpcfgm_model.N_PULSES = cfg['n_pulses']

        from ejcpcfgm_model import run_ejcpcfgm
        result = run_ejcpcfgm(
            list(train_data), list(test_data),
            pop_size=20, max_iter=1500,
            ridge_lambda=cfg['ridge_lambda'],
            theta_lb=cfg['theta_lb'], theta_ub=0.0,
            sigma_lb=0.0, sigma_ub=1.0,
            cq_lb=0.0, cq_ub=cfg['cq_ub'],
            r_lb=0.0, r_ub=1.0
        )

        return {
            'train_predictions': np.array(result.get('train_predictions', [])),
            'test_predictions': np.array(result.get('test_predictions', [])),
            'success': True
        }
    except Exception:
        return {'success': False}


# ==================== 运行所有模型 ====================
def run_all_models_for_dataset(ds_name, ds_info):
    train_data = np.array(ds_info['train'])
    test_data = np.array(ds_info['test'])
    n_forecast = len(test_data)
    train_years = ds_info['train_years']
    test_years = ds_info['test_years']
    n_train = len(train_data)

    results = {}

    cfgm_cfg = EJCPCFGM_BEST[ds_name]
    cfgm_res = cfgm_predict(train_data, test_data, cfgm_cfg)
    if cfgm_res['success']:
        train_preds_no_first = cfgm_res['train_predictions'][1:]
        results['EJCPCFGM'] = {
            'fitted': train_preds_no_first,
            'forecast': cfgm_res['test_predictions'],
            'ws_offset': 1
        }

    res = gm11_predict(train_data, n_forecast)
    if res['success']:
        results['GM(1,1)'] = {'fitted': res['fitted'][1:], 'forecast': res['forecast'], 'ws_offset': 1}

    res = fgm_predict(train_data, n_forecast)
    if res['success']:
        results['FGM'] = {'fitted': res['fitted'][1:], 'forecast': res['forecast'], 'ws_offset': 1}

    res = jfgm_predict(train_data, n_forecast)
    if res['success']:
        results['JFGM'] = {'fitted': res['fitted'][1:], 'forecast': res['forecast'], 'ws_offset': 1}

    res = ngbm_predict(train_data, n_forecast)
    if res['success']:
        results['NGBM'] = {'fitted': res['fitted'][1:], 'forecast': res['forecast'], 'ws_offset': 1}

    res = arima_predict(train_data, n_forecast)
    if res['success']:
        fitted_no_first = res['fitted'][1:]
        results['ARIMA'] = {'fitted': fitted_no_first, 'forecast': res['forecast'], 'ws_offset': 1}

    lssvr_fitted, lssvr_forecast, lssvr_ws = load_lssvr_results(ds_name)
    if lssvr_fitted is not None:
        results['LSSVR'] = {'fitted': lssvr_fitted, 'forecast': lssvr_forecast, 'ws_offset': lssvr_ws}

    lstm_fitted, lstm_forecast, lstm_ws = load_lstm_results(ds_name, train_data, test_data)
    if lstm_fitted is not None:
        results['LSTM'] = {'fitted': lstm_fitted, 'forecast': lstm_forecast, 'ws_offset': lstm_ws}

    mape_results = {}
    for method, data in results.items():
        if 'train_mape' in data and 'test_mape' in data:
            mape_results[method] = dict(train_mape=data['train_mape'], test_mape=data['test_mape'])
            continue

        fitted = data['fitted']
        forecast = data['forecast']
        ws_off = data['ws_offset']

        train_actual = train_data[ws_off:] if ws_off > 0 else train_data
        train_fitted = fitted
        train_mape = calculate_mape(train_actual, train_fitted) if len(train_actual) > 0 else float('nan')
        test_mape = calculate_mape(test_data, forecast)

        mape_results[method] = dict(train_mape=train_mape, test_mape=test_mape)

    return results, mape_results, train_years, test_years


# ==================== 生成 Excel ====================
def create_excel(all_results, output_path):
    wb = Workbook()

    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    subheader_font = Font(name='Arial', bold=True, size=10)
    subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    best_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    ml_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # ===== Sheet 1: MAPE汇总 =====
    ws_summary = wb.active
    ws_summary.title = 'MAPE_Summary'

    ws_summary.merge_cells('A1:I1')
    ws_summary['A1'] = '8 Models MAPE Comparison (Test MAPE %)'
    ws_summary['A1'].font = Font(name='Arial', bold=True, size=14)
    ws_summary['A1'].alignment = center_align

    headers = ['Dataset', 'EJCPCFGM', 'ARIMA', 'GM(1,1)', 'FGM', 'JFGM', 'NGBM', 'LSSVR', 'LSTM']
    for col, h in enumerate(headers, 1):
        c = ws_summary.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = thin_border

    row = 4
    for ds_name, (_, mape_results, _, _) in all_results.items():
        ws_summary.cell(row=row, column=1, value=ds_name).font = subheader_font
        ws_summary.cell(row=row, column=1).border = thin_border
        ws_summary.cell(row=row, column=1).alignment = center_align

        test_mapes = {m: mape_results[m]['test_mape'] for m in mape_results if not np.isnan(mape_results[m]['test_mape'])}
        best_model = min(test_mapes, key=test_mapes.get) if test_mapes else ''

        for col, method in enumerate(['EJCPCFGM', 'ARIMA', 'GM(1,1)', 'FGM', 'JFGM', 'NGBM', 'LSSVR', 'LSTM'], 2):
            c = ws_summary.cell(row=row, column=col)
            if method in mape_results and not np.isnan(mape_results[method]['test_mape']):
                c.value = f"{mape_results[method]['test_mape']:.4f}%"
            else:
                c.value = 'N/A'
            c.alignment = center_align
            c.border = thin_border
            c.font = data_font
            if method == best_model:
                c.fill = best_fill
                c.font = Font(name='Arial', size=10, bold=True)
        row += 1

    ws_summary.column_dimensions['A'].width = 25
    for col in 'BCDEFGHI':
        ws_summary.column_dimensions[col].width = 13

    # ===== 每个数据集的详细Sheet =====
    for ds_name, (pred_results, mape_results, train_years, test_years) in all_results.items():
        short_name = ds_name.replace('_', ' ')[:31]
        ws = wb.create_sheet(title=short_name)

        methods = list(pred_results.keys())
        n_train = len(train_years)
        n_test = len(test_years)

        train_actual = list(DATASETS[ds_name]['train'])
        test_actual = list(DATASETS[ds_name]['test'])

        # 标题
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(methods) * 2 + 2)
        ws['A1'] = f'{ds_name} - All Models Predictions'
        ws['A1'].font = Font(name='Arial', bold=True, size=14)
        ws['A1'].alignment = center_align

        # 列头
        header_row = 3
        col = 1
        for h, f in [('Year', header_font), ('Type', header_font), ('Actual', header_font)]:
            ws.cell(row=header_row, column=col, value=h).font = f
            ws.cell(row=header_row, column=col).fill = header_fill
            ws.cell(row=header_row, column=col).alignment = center_align
            ws.cell(row=header_row, column=col).border = thin_border
            col += 1

        for method in methods:
            is_ml = method in ['LSSVR', 'LSTM']
            fnt = Font(name='Arial', bold=True, size=10, color='9C6500') if is_ml else subheader_font
            fill = ml_fill if is_ml else subheader_fill
            ws.cell(row=header_row, column=col, value=method).font = fnt
            ws.cell(row=header_row, column=col).fill = fill
            ws.cell(row=header_row, column=col).alignment = center_align
            ws.cell(row=header_row, column=col).border = thin_border
            ws.cell(row=header_row, column=col+1, value='APE%').font = fnt
            ws.cell(row=header_row, column=col+1).fill = fill
            ws.cell(row=header_row, column=col+1).alignment = center_align
            ws.cell(row=header_row, column=col+1).border = thin_border
            col += 2

        ws.cell(row=header_row, column=col, value='Train MAPE').font = header_font
        ws.cell(row=header_row, column=col).fill = header_fill
        ws.cell(row=header_row, column=col).alignment = center_align
        ws.cell(row=header_row, column=col).border = thin_border
        col += 1
        ws.cell(row=header_row, column=col, value='Test MAPE').font = header_font
        ws.cell(row=header_row, column=col).fill = header_fill
        ws.cell(row=header_row, column=col).alignment = center_align
        ws.cell(row=header_row, column=col).border = thin_border

        # 写训练数据
        for i, year in enumerate(train_years):
            r = header_row + 1 + i
            col = 1
            actual_val = train_actual[i]

            ws.cell(row=r, column=col, value=year).font = data_font
            ws.cell(row=r, column=col).alignment = center_align
            ws.cell(row=r, column=col).border = thin_border
            col += 1
            ws.cell(row=r, column=col, value='Train').font = data_font
            ws.cell(row=r, column=col).alignment = center_align
            ws.cell(row=r, column=col).border = thin_border
            col += 1
            ws.cell(row=r, column=col, value=actual_val).font = data_font
            ws.cell(row=r, column=col).alignment = right_align
            ws.cell(row=r, column=col).border = thin_border
            ws.cell(row=r, column=col).number_format = '0.00'
            col += 1

            for method in methods:
                ws_off = pred_results[method]['ws_offset']
                fitted = pred_results[method]['fitted']

                if i < ws_off:
                    for _ in range(2):
                        c = ws.cell(row=r, column=col)
                        c.value = '-'
                        c.font = Font(name='Arial', size=10, color='AAAAAA')
                        c.alignment = center_align
                        c.border = thin_border
                        col += 1
                else:
                    fit_idx = i - ws_off
                    if fit_idx < len(fitted):
                        pred_val = fitted[fit_idx]
                        c = ws.cell(row=r, column=col)
                        c.value = pred_val
                        c.font = data_font
                        c.alignment = right_align
                        c.border = thin_border
                        c.number_format = '0.00'
                        col += 1

                        c = ws.cell(row=r, column=col)
                        if actual_val != 0:
                            ape = abs((actual_val - pred_val) / actual_val) * 100
                            c.value = f"{ape:.2f}%"
                        else:
                            c.value = '-'
                        c.font = data_font
                        c.alignment = center_align
                        c.border = thin_border
                        col += 1
                    else:
                        for _ in range(2):
                            c = ws.cell(row=r, column=col)
                            c.value = '-'
                            c.font = Font(name='Arial', size=10, color='AAAAAA')
                            c.alignment = center_align
                            c.border = thin_border
                            col += 1

            # MAPE列留空
            ws.cell(row=r, column=col, value='').border = thin_border
            col += 1
            ws.cell(row=r, column=col, value='').border = thin_border

        # 写测试数据
        for i, year in enumerate(test_years):
            r = header_row + 1 + n_train + i
            col = 1
            actual_val = test_actual[i]

            ws.cell(row=r, column=col, value=year).font = Font(name='Arial', size=10, bold=True)
            ws.cell(row=r, column=col).alignment = center_align
            ws.cell(row=r, column=col).border = thin_border
            ws.cell(row=r, column=col).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            col += 1
            ws.cell(row=r, column=col, value='Test').font = Font(name='Arial', size=10, bold=True, color='C00000')
            ws.cell(row=r, column=col).alignment = center_align
            ws.cell(row=r, column=col).border = thin_border
            ws.cell(row=r, column=col).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            col += 1
            ws.cell(row=r, column=col, value=actual_val).font = Font(name='Arial', size=10, bold=True)
            ws.cell(row=r, column=col).alignment = right_align
            ws.cell(row=r, column=col).border = thin_border
            ws.cell(row=r, column=col).number_format = '0.00'
            ws.cell(row=r, column=col).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            col += 1

            for method in methods:
                forecast = pred_results[method]['forecast']
                if i < len(forecast):
                    pred_val = forecast[i]
                    c = ws.cell(row=r, column=col)
                    c.value = pred_val
                    c.font = Font(name='Arial', size=10, bold=True)
                    c.alignment = right_align
                    c.border = thin_border
                    c.number_format = '0.00'
                    c.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    col += 1

                    c = ws.cell(row=r, column=col)
                    if actual_val != 0:
                        ape = abs((actual_val - pred_val) / actual_val) * 100
                        c.value = f"{ape:.2f}%"
                    else:
                        c.value = '-'
                    c.font = Font(name='Arial', size=10, bold=True)
                    c.alignment = center_align
                    c.border = thin_border
                    c.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    col += 1
                else:
                    for _ in range(2):
                        c = ws.cell(row=r, column=col)
                        c.value = '-'
                        c.font = Font(name='Arial', size=10, color='AAAAAA')
                        c.alignment = center_align
                        c.border = thin_border
                        col += 1

            ws.cell(row=r, column=col, value='').border = thin_border
            col += 1
            ws.cell(row=r, column=col, value='').border = thin_border

        # 写 MAPE 汇总行
        mape_row = header_row + 1 + n_train + n_test + 2

        ws.cell(row=mape_row, column=1, value='Train MAPE').font = subheader_font
        ws.cell(row=mape_row, column=1).fill = subheader_fill
        ws.cell(row=mape_row, column=1).border = thin_border
        ws.cell(row=mape_row, column=2, value='').fill = subheader_fill
        ws.cell(row=mape_row, column=2).border = thin_border
        ws.cell(row=mape_row, column=3, value='').fill = subheader_fill
        ws.cell(row=mape_row, column=3).border = thin_border
        col = 4
        for method in methods:
            c = ws.cell(row=mape_row, column=col)
            if method in mape_results and not np.isnan(mape_results[method]['train_mape']):
                c.value = f"{mape_results[method]['train_mape']:.4f}%"
            else:
                c.value = 'N/A'
            c.font = subheader_font
            c.fill = subheader_fill
            c.alignment = center_align
            c.border = thin_border
            ws.cell(row=mape_row, column=col+1, value='').fill = subheader_fill
            ws.cell(row=mape_row, column=col+1).border = thin_border
            col += 2
        ws.cell(row=mape_row, column=col, value='').fill = subheader_fill
        ws.cell(row=mape_row, column=col).border = thin_border
        ws.cell(row=mape_row, column=col+1, value='').fill = subheader_fill
        ws.cell(row=mape_row, column=col+1).border = thin_border

        mape_row += 1
        test_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        ws.cell(row=mape_row, column=1, value='Test MAPE').font = Font(name='Arial', size=10, bold=True, color='C00000')
        ws.cell(row=mape_row, column=1).fill = test_fill
        ws.cell(row=mape_row, column=1).border = thin_border
        ws.cell(row=mape_row, column=2, value='').fill = test_fill
        ws.cell(row=mape_row, column=2).border = thin_border
        ws.cell(row=mape_row, column=3, value='').fill = test_fill
        ws.cell(row=mape_row, column=3).border = thin_border
        col = 4
        for method in methods:
            c = ws.cell(row=mape_row, column=col)
            if method in mape_results and not np.isnan(mape_results[method]['test_mape']):
                c.value = f"{mape_results[method]['test_mape']:.4f}%"
            else:
                c.value = 'N/A'
            c.font = Font(name='Arial', size=10, bold=True, color='C00000')
            c.fill = test_fill
            c.alignment = center_align
            c.border = thin_border
            ws.cell(row=mape_row, column=col+1, value='').fill = test_fill
            ws.cell(row=mape_row, column=col+1).border = thin_border
            col += 2
        ws.cell(row=mape_row, column=col, value='').fill = test_fill
        ws.cell(row=mape_row, column=col).border = thin_border
        ws.cell(row=mape_row, column=col+1, value='').fill = test_fill
        ws.cell(row=mape_row, column=col+1).border = thin_border

        # 列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 14
        for i in range(len(methods)):
            cl = chr(ord('D') + i * 2)
            ws.column_dimensions[cl].width = 14
            ws.column_dimensions[chr(ord(cl) + 1)].width = 10

    wb.save(output_path)
    print(f"\nExcel saved: {output_path}")


def main():
    all_results = {}

    for ds_name, ds_info in DATASETS.items():
        print(f"[{ds_name}]")
        pred_results, mape_results, train_years, test_years = run_all_models_for_dataset(ds_name, ds_info)
        all_results[ds_name] = (pred_results, mape_results, train_years, test_years)

        for method in sorted(mape_results.keys()):
            tm = mape_results[method]['train_mape']
            pm = mape_results[method]['test_mape']
            print(f"  {method:12s}  train={tm:7.4f}%  test={pm:7.4f}%")

    output_path = os.path.join(SCRIPT_DIR, 'All_8_Models_Summary.xlsx')
    create_excel(all_results, output_path)
    print(f"saved: {output_path}")


if __name__ == "__main__":
    main()
